#!/usr/bin/env python3
"""Tool to add vocabulary entries to Vocabulary.xlsx"""

import sys
import openpyxl
from pathlib import Path

EXCEL_PATH = Path("/home/icass/rockit/Vocabulary.xlsx")


def add_vocabulary(key: str, english_text: str) -> None:
    """Add a new vocabulary entry to the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Find the table and extend it
    table = next(iter(ws.tables.values()))
    from openpyxl.utils import range_boundaries, get_column_letter
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    next_row = max_row + 1

    ws.cell(row=next_row, column=1, value=key)
    ws.cell(row=next_row, column=2, value=english_text)

    for col in range(3, max_col + 1):
        formula = f"=_xlfn.TRANSLATE($B{next_row},$B$2,{get_column_letter(col)}$2)"
        ws.cell(row=next_row, column=col, value=formula)

    # Extend table ref to include new row
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{next_row}"

    wb.save(EXCEL_PATH)
    print(f"Added: {key} = {english_text}")


def main():
    args = sys.argv[1:]
    if len(args) < 2 or len(args) % 2 != 0:
        print("Usage: python add_vocabulary.py <key1> <english_text1> [<key2> <english_text2> ...]")
        print("Example: python add_vocabulary.py ADMIN_BUILDS_TITLE 'Android Builds' ADMIN_CANCEL 'Cancel'")
        sys.exit(1)

    pairs = [(args[i], args[i + 1]) for i in range(0, len(args), 2)]
    for key, english_text in pairs:
        add_vocabulary(key, english_text)


if __name__ == "__main__":
    main()
