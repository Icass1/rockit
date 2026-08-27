import json
import os
import shutil
import sqlite3
import sys
import uuid
import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple
from concurrent.futures import ThreadPoolExecutor
import argparse

from backend.utils.logger import getLogger

logger = getLogger(__name__)

parser = argparse.ArgumentParser()
parser.add_argument("--env", type=str, help="Path to env file to load")
parser.add_argument(
    "--lrc-lib-path",
    type=str,
    default="../lrclib-db-dump-20260519T172012Z.sqlite3",
    help="Path to LRCLIB SQLite dump",
)
parser.add_argument(
    "--db",
    type=str,
    default="database.db",
    help="Path to the legacy SQLite database.db",
)
parser.add_argument("command", nargs="?", help="Command to run")
args, _ = parser.parse_known_args()
command_to_run = args.command if args.command else ""

if args.env:
    if not os.path.exists(args.env):
        print(f"Env file not found: {args.env}")
        sys.exit(1)
    print(f"Loading {args.env}...")
    os.environ["ROCKIT_ENV_FILE"] = args.env


async def import_vocabulary() -> None:
    """Import vocabulary from Vocabulary.xlsx file."""
    from backend.core.access.db import rockit_db

    types_path = Path("frontend/packages/shared/src/models/types/vocabulary.ts")

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required. Install with: pip install openpyxl")
        return

    file_path = os.path.join("Vocabulary.xlsx")

    if not os.path.exists(file_path):
        logger.error(f"Vocabulary.xlsx not found at {file_path}")
        return

    logger.info(f"Reading vocabulary from {file_path}")

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    if sheet is None:
        logger.error("No active sheet found in workbook")
        workbook.close()
        return

    first_row: List[Any] = [cell.value for cell in sheet[1]]
    logger.info(f"Found columns: {first_row}")

    language_columns: List[str] = [str(h) for h in first_row[1:] if h is not None]
    logger.info(f"Found languages: {language_columns}")

    second_row: List[Any] = [cell.value for cell in sheet[2]]

    if not second_row or second_row[0] is None:
        logger.error("First column must be 'KEY'")
        return

    if second_row[0] != "KEY":
        logger.error("First column must be 'KEY'")
        return

    language_code_columns: List[str] = [str(h) for h in second_row[1:] if h is not None]
    logger.info(f"Found languages codes: {language_code_columns}")

    types_content: List[str] = [
        "// This file is generated using: python3 -m backend import_vocabulary",
        "// Do not modify this file manually.",
        "",
        "export interface Vocabulary {",
    ]

    vocabulary_data: Dict[str, Dict[str, str]] = {
        lang: {} for lang in language_code_columns
    }
    all_keys: List[str] = []

    max_row: int = sheet.max_row or 0
    for row_idx in range(3, max_row + 1):
        key_cell = sheet.cell(row=row_idx, column=1).value
        if not key_cell:
            continue

        if type(key_cell) != str:
            logger.warning(f"{key_cell} is not a number")
            continue

        all_keys.append(str(key_cell))

        if key_cell.startswith(" ") or key_cell.endswith(" "):
            logger.critical(f"Traling spaces detected in key '{key_cell}'")

        types_content.append(f"    {key_cell}: string;")

        for col_idx, lang_code in enumerate(language_code_columns, start=2):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value:
                vocabulary_data[lang_code][str(key_cell)] = str(value)

    types_content.append("}")
    types_content.append("")
    types_path.write_text("\n".join(types_content))

    logger.info(f"Found {len(all_keys)} vocabulary keys")

    async with rockit_db.session_scope_async() as session:
        from backend.core.framework.language import Language
        from backend.core.framework.vocabulary import Vocabulary
        from backend.core.framework.models.vocabulary import VocabularyImportData

        for lang_name, lang_code in zip(language_columns, language_code_columns):
            a_result = await Language.get_or_create_language(
                session=session, lang_code=lang_code, language=lang_name
            )
            if a_result.is_ok():
                logger.info(f"Language {lang_code} ready")
            else:
                logger.error(f"Error with language {lang_code}: {a_result.message()}")

        import_data = VocabularyImportData(vocabulary=vocabulary_data)
        a_result_import = await Vocabulary.import_vocabulary_from_dict(
            session=session, vocabulary_data=import_data
        )
        if a_result_import.is_ok():
            logger.info("Vocabulary imported successfully")
        else:
            logger.error(f"Error importing vocabulary: {a_result_import.message()}")

        a_result_cleanup = await Vocabulary.remove_keys_not_in_import(
            session=session, valid_keys=all_keys
        )
        if a_result_cleanup.is_ok():
            logger.info("Cleanup completed")
        else:
            logger.error(f"Error during cleanup: {a_result_cleanup.message()}")

    workbook.close()
    logger.info("Import complete!")


async def fix_shared_images_async(sqlite_path: str) -> None:
    """Fix images referenced by more than one media row.

    Images referenced by both a Spotify media and a non-Spotify media (e.g. a
    YouTube channel/video) keep the non-Spotify reference and get a brand new
    image for each Spotify media. Images referenced by two Spotify medias keep
    one image and get a new image for each of the others. The original Spotify
    image URLs are looked up in the legacy SQLite database.db.
    """

    import requests as req

    from sqlalchemy import select, update

    from backend.constants import IMAGES_PATH
    from backend.utils.colorExtractor import extract_dominant_color

    from backend.core.aResult import AResult
    from backend.core.access.db import rockit_db
    from backend.core.access.imageAccess import ImageAccess
    from backend.core.access.db.ormModels.image import ImageRow
    from backend.core.access.db.ormModels.user import UserRow

    from backend.default.access.db.ormModels.playlist import (
        PlaylistRow as DefaultPlaylistRow,
    )

    from backend.rockit.access.db.ormModels.album import RockitAlbumRow
    from backend.rockit.access.db.ormModels.artist import RockitArtistRow
    from backend.rockit.access.db.ormModels.song import RockitSongRow
    from backend.rockit.access.db.ormModels.video import RockitVideoRow

    from backend.spotify.access.db.ormModels.album import AlbumRow as SpotifyAlbumRow
    from backend.spotify.access.db.ormModels.artist import (
        ArtistRow as SpotifyArtistRow,
    )
    from backend.spotify.access.db.ormModels.playlist import (
        PlaylistRow as SpotifyPlaylistRow,
    )

    from backend.spotifyScrapper.access.db.ormModels.album import (
        AlbumRow as ScrapperAlbumRow,
    )
    from backend.spotifyScrapper.access.db.ormModels.artist import (
        ArtistRow as ScrapperArtistRow,
    )
    from backend.spotifyScrapper.access.db.ormModels.playlist import (
        PlaylistRow as ScrapperPlaylistRow,
    )

    from backend.youtube.access.db.ormModels.channel import ChannelRow
    from backend.youtube.access.db.ormModels.playlist import YoutubePlaylistRow
    from backend.youtube.access.db.ormModels.video import VideoRow as YoutubeVideoRow

    from backend.youtubeMusic.access.db.ormModels.album import (
        AlbumRow as YoutubeMusicAlbumRow,
    )
    from backend.youtubeMusic.access.db.ormModels.artist import (
        ArtistRow as YoutubeMusicArtistRow,
    )
    from backend.youtubeMusic.access.db.ormModels.playlist import (
        PlaylistRow as YoutubeMusicPlaylistRow,
    )
    from backend.youtubeMusic.access.db.ormModels.track import (
        TrackRow as YoutubeMusicTrackRow,
    )

    if not os.path.exists(sqlite_path):
        logger.error(f"SQLite database not found: {sqlite_path}")
        return

    @dataclass
    class ImageReference:
        """Reference of a media row to an image."""

        label: str
        image_id: int
        media_id: int
        model: Any
        spotify_id: str | None = None
        image_folder: str | None = None
        sqlite_table: str | None = None

    spotify_tables: List[Tuple[Any, str, str, str]] = [
        # (model, label, image folder for new paths, sqlite table for url lookup)
        (SpotifyAlbumRow, "spotify album", "albums", "album"),
        (SpotifyArtistRow, "spotify artist", "artists", "artist"),
        (SpotifyPlaylistRow, "spotify playlist", "playlists", "playlist"),
        (ScrapperAlbumRow, "spotify_scrapper album", "albums", "album"),
        (ScrapperArtistRow, "spotify_scrapper artist", "artists", "artist"),
        (ScrapperPlaylistRow, "spotify_scrapper playlist", "playlists", "playlist"),
    ]

    keep_tables: List[Tuple[Any, str]] = [
        (UserRow, "core user"),
        (DefaultPlaylistRow, "default playlist"),
        (RockitAlbumRow, "rockit album"),
        (RockitArtistRow, "rockit artist"),
        (RockitSongRow, "rockit song"),
        (RockitVideoRow, "rockit video"),
        (ChannelRow, "youtube channel"),
        (YoutubeVideoRow, "youtube video"),
        (YoutubePlaylistRow, "youtube playlist"),
        (YoutubeMusicAlbumRow, "youtube_music album"),
        (YoutubeMusicArtistRow, "youtube_music artist"),
        (YoutubeMusicPlaylistRow, "youtube_music playlist"),
        (YoutubeMusicTrackRow, "youtube_music track"),
    ]

    def get_spotify_image_url(
        sqlite_conn: sqlite3.Connection,
        sqlite_table: str,
        spotify_id: str,
    ) -> str | None:
        """Get the first Spotify CDN image URL from database.db for a media id."""

        try:
            cursor = sqlite_conn.execute(
                f'SELECT images FROM "{sqlite_table}" WHERE id = ?',
                (spotify_id,),
            )
            row = cursor.fetchone()
            if row is None or not row[0]:
                return None
            images_data = json.loads(row[0])
            if not images_data:
                return None
            url = images_data[0].get("url")
            return url if isinstance(url, str) and url else None
        except Exception as e:
            logger.error(
                f"Error getting Spotify image url for {sqlite_table}/{spotify_id}: {e}"
            )
            return None

    async def create_spotify_image(
        session: Any,
        sqlite_conn: sqlite3.Connection,
        original_image: ImageRow,
        existing_paths: set[str],
        ref: ImageReference,
    ) -> ImageRow | None:
        """Create a new image row for a Spotify media and download its file."""

        spotify_id: str | None = ref.spotify_id
        if not spotify_id:
            logger.error(f"Skipping {ref.label} media {ref.media_id}: no spotify id")
            return None

        url: str | None = None
        if ref.sqlite_table:
            url = get_spotify_image_url(sqlite_conn, ref.sqlite_table, spotify_id)

        new_path: str = f"spotify/{ref.image_folder}/{spotify_id}.png"
        if new_path in existing_paths:
            new_path = (
                f"spotify/{ref.image_folder}/{spotify_id}-{uuid.uuid4().hex[:8]}.png"
            )

        full_path: str = os.path.join(IMAGES_PATH, new_path)

        if url and not os.path.exists(full_path):
            try:
                response = req.get(url, timeout=30)
                if response.status_code == 200:
                    os.makedirs(os.path.dirname(full_path), exist_ok=True)
                    with open(full_path, "wb") as f:
                        f.write(response.content)
                else:
                    logger.warning(
                        f"Download failed for {url}: status {response.status_code}"
                    )
            except Exception as e:
                logger.error(f"Error downloading image {url}: {e}")

        if not os.path.exists(full_path):
            original_full_path: str = os.path.join(IMAGES_PATH, original_image.path)
            if os.path.exists(original_full_path):
                try:
                    os.makedirs(os.path.dirname(full_path), exist_ok=True)
                    shutil.copy2(original_full_path, full_path)
                    logger.info(f"Copied {original_image.path} to {new_path}")
                except Exception as e:
                    logger.error(f"Error copying image to {new_path}: {e}")

        dominant_color: str = original_image.dominant_color
        extracted: str | None = await extract_dominant_color(full_path)
        if extracted:
            dominant_color = extracted

        a_result: AResult[ImageRow] = await ImageAccess.insert_image_async(
            session=session,
            path=new_path,
            url=url,
            dominant_color=dominant_color,
        )
        if a_result.is_not_ok():
            logger.error(
                f"Error creating image for {ref.label} {spotify_id}: {a_result.message()}"
            )
            return None

        new_image: ImageRow = a_result.result()
        logger.info(
            f"Created image id={new_image.id} path={new_image.path} url={url} "
            f"for {ref.label} {spotify_id}"
        )
        return new_image

    sqlite_conn = sqlite3.connect(sqlite_path)

    try:
        async with rockit_db.session_scope_async() as session:
            keep_image_ids: set[int] = set()
            spotify_refs: Dict[int, List[ImageReference]] = {}

            for model, _label in keep_tables:
                result = await session.execute(select(model.id, model.image_id))
                for row in result.all():
                    if row.image_id is not None:
                        keep_image_ids.add(row.image_id)

            existing_paths: set[str] = set()
            a_result_images: AResult[List[ImageRow]] = (
                await ImageAccess.get_all_images_async(session=session)
            )
            if a_result_images.is_ok():
                existing_paths = {img.path for img in a_result_images.result()}

            for model, label, image_folder, sqlite_table in spotify_tables:
                result = await session.execute(
                    select(model.id, model.image_id, model.spotify_id)
                )
                for row in result.all():
                    if row.image_id is None:
                        continue
                    spotify_refs.setdefault(row.image_id, []).append(
                        ImageReference(
                            label=label,
                            image_id=row.image_id,
                            media_id=row.id,
                            model=model,
                            spotify_id=row.spotify_id,
                            image_folder=image_folder,
                            sqlite_table=sqlite_table,
                        )
                    )

            candidate_ids: List[int] = [
                image_id
                for image_id, refs in spotify_refs.items()
                if len(refs) + (1 if image_id in keep_image_ids else 0) > 1
            ]

            logger.info(f"Found {len(candidate_ids)} shared images to fix")
            for image_id in sorted(candidate_ids):
                labels = ", ".join(ref.label for ref in spotify_refs[image_id])
                logger.info(f"  - image {image_id}: {labels}")

            processed_count: int = 0
            created_count: int = 0
            error_count: int = 0

            for image_id in sorted(candidate_ids):
                refs: List[ImageReference] = spotify_refs[image_id]
                has_keep: bool = image_id in keep_image_ids

                a_result_image: AResult[ImageRow] = (
                    await ImageAccess.get_image_from_id_async(
                        session=session, id=image_id
                    )
                )
                if a_result_image.is_not_ok():
                    logger.error(
                        f"Image {image_id} not found, skipping. {a_result_image.message()}"
                    )
                    error_count += 1
                    continue
                original_image: ImageRow = a_result_image.result()

                refs_to_keep: List[ImageReference] = []
                refs_to_split: List[ImageReference] = refs

                if not has_keep:
                    keep_ref: ImageReference | None = None
                    for ref in refs:
                        if (
                            ref.image_folder
                            and ref.spotify_id
                            and original_image.path
                            == f"spotify/{ref.image_folder}/{ref.spotify_id}.png"
                        ):
                            keep_ref = ref
                            break
                    if keep_ref is None:
                        keep_ref = refs[0]
                    refs_to_keep = [keep_ref]
                    refs_to_split = [r for r in refs if r is not keep_ref]

                for ref in refs_to_keep:
                    if has_keep:
                        logger.info(
                            f"Image {image_id} (path={original_image.path}) "
                            f"kept for non-Spotify media"
                        )
                        continue
                    if ref.spotify_id and ref.sqlite_table:
                        url = get_spotify_image_url(
                            sqlite_conn, ref.sqlite_table, ref.spotify_id
                        )
                        if url and url != original_image.url:
                            original_image.url = url
                            logger.info(
                                f"Image {image_id} kept for {ref.label} "
                                f"(media {ref.media_id}), url set to {url}"
                            )
                    else:
                        logger.info(
                            f"Image {image_id} kept for {ref.label} (media {ref.media_id})"
                        )

                for ref in refs_to_split:
                    new_image = await create_spotify_image(
                        session=session,
                        sqlite_conn=sqlite_conn,
                        original_image=original_image,
                        existing_paths=existing_paths,
                        ref=ref,
                    )
                    if new_image is None:
                        error_count += 1
                        continue
                    existing_paths.add(new_image.path)
                    await session.execute(
                        update(ref.model)
                        .where(ref.model.id == ref.media_id)
                        .values(image_id=new_image.id)
                    )
                    logger.info(
                        f"Image id={ref.image_id} of {ref.label} media "
                        f"{ref.media_id} updated to image id={new_image.id}"
                    )
                    created_count += 1

                if refs_to_split:
                    processed_count += 1
                await session.commit()

            logger.info(
                f"Fixed {processed_count} shared images, created {created_count} "
                f"new images, {error_count} errors"
            )
    finally:
        sqlite_conn.close()


async def backfill_dominant_colors() -> None:
    """Extract and store dominant_color for all images that are missing it."""
    import os

    from backend.constants import IMAGES_PATH
    from backend.core.access.db import rockit_db
    from backend.core.access.imageAccess import ImageAccess

    async with rockit_db.session_scope_async() as session:
        a_result = await ImageAccess.get_images_needing_color_backfill_async(
            session=session
        )
        if a_result.is_not_ok():
            logger.error(f"Error fetching images for backfill. {a_result.info()}")
            return

        images = a_result.result()

        if not images:
            logger.info("No images need dominant_color backfill")
            return

        logger.info(f"Backfilling dominant_color for {len(images)} images...")

        from backend.utils.colorExtractor import extract_dominant_color

        for index, image in enumerate(images):
            image_path = os.path.join(IMAGES_PATH, image.path)
            color = await extract_dominant_color(image_path)
            if color is not None:
                await ImageAccess.update_image_dominant_color_async(
                    session=session, image=image, dominant_color=color
                )
                logger.info(f"Backfilled {image.path} -> {color}")
            else:
                logger.warning(f"Error backfilling {image.path}")

            if index % 100 == 0:
                await session.commit()

        await session.commit()

        logger.info("Dominant color backfill complete")


async def main() -> None:
    from backend.core.access.db import rockit_db
    from backend.core import add_initial_content_async

    # Only init DB for commands that need it.
    needs_db: bool = not command_to_run in ["models"]
    if needs_db:
        try:
            await rockit_db.async_init()
        except Exception as e:
            logger.critical(f"Error initializing database: {e}")
            sys.exit()

        await rockit_db.wait_for_session_local_async()

    first_loop = True

    def run_input(prompt: str) -> str:
        return input(prompt)

    with ThreadPoolExecutor(max_workers=1) as executor:
        while True:
            command: str
            if first_loop and command_to_run != "":
                command = command_to_run
                first_loop = False
            else:
                try:
                    future = executor.submit(run_input, "> ")
                    command = future.result()
                except KeyboardInterrupt:
                    break

            if command == "exit":
                break

            elif command == "reinit":
                await rockit_db.reinit()

            elif command == "models":
                from backend.utils.zodGenerator import generate_zod_schemas
                from backend.utils.httpMethodsGenerator import http_methods_generator

                await generate_zod_schemas()
                await http_methods_generator()

            elif command == "import-vocabulary":
                await import_vocabulary()

            elif command == "init-db":
                if hasattr(rockit_db, "engine"):
                    await rockit_db.engine.dispose()
                await rockit_db.async_init()

                await add_initial_content_async()

                logger.info("Database initialized")

            elif command == "backfill-dominant-colors":
                await backfill_dominant_colors()

            elif command == "cleanup-images":
                from backend.core.access.imageAccess import ImageAccess
                from backend.constants import IMAGES_PATH

                db_paths: set[str] = set()
                fs_paths: set[str] = set()

                async with rockit_db.session_scope_async() as session:
                    a_result = await ImageAccess.get_all_images_async(session=session)
                    if a_result.is_not_ok():
                        logger.error(f"Error getting images: {a_result.message()}")
                        continue

                    for img in a_result.result():
                        db_paths.add(img.path)

                logger.info(f"Found {len(db_paths)} images in database")

                for root, _, files in os.walk(IMAGES_PATH):
                    for file in files:
                        rel_path = os.path.relpath(
                            os.path.join(root, file), IMAGES_PATH
                        )
                        fs_paths.add(rel_path)

                logger.info(f"Found {len(fs_paths)} images in filesystem")

                extra_files = fs_paths - db_paths

                if not extra_files:
                    logger.info("No extra images to clean up")
                    continue

                logger.info(f"Extra images ({len(extra_files)}):")
                for path in sorted(extra_files):
                    logger.info(f"  - {path}")

                confirm = input("Delete these files? This cannot be undone (y/N): ")
                if confirm.lower() != "y":
                    logger.info("Cancelled")
                    continue

                deleted_count = 0
                for rel_path in extra_files:
                    full_path = os.path.join(IMAGES_PATH, rel_path)
                    try:
                        os.remove(full_path)
                        logger.info(f"Deleted: {rel_path}")
                        deleted_count += 1
                    except Exception as e:
                        logger.error(f"Error deleting {rel_path}: {e}")

                logger.info(f"Deleted {deleted_count} files")

            elif command == "fix-images":
                import requests as req

                from backend.core.access.imageAccess import ImageAccess
                from backend.constants import IMAGES_PATH

                async with rockit_db.session_scope_async() as session:
                    a_result = await ImageAccess.get_all_images_async(session=session)
                    if a_result.is_not_ok():
                        logger.error(f"Error getting images: {a_result.message()}")
                        continue

                    images = a_result.result()
                    logger.info(f"Found {len(images)} images in database")

                    missing_count = 0
                    fixed_count = 0
                    error_count = 0
                    skipped_count = 0

                    for image in images:
                        if image.path.startswith("/"):
                            logger.error(
                                f"Image path ({image.path}) starts with /, modify it in database."
                            )
                            error_count += 1
                            continue

                        full_path = os.path.join(IMAGES_PATH, image.path)
                        if os.path.exists(full_path):
                            continue

                        missing_count += 1
                        if not image.url:
                            logger.warning(
                                f"Image {image.path} is missing and has no url to re-download"
                            )
                            skipped_count += 1
                            continue

                        try:
                            response = req.get(image.url, timeout=30)
                            if response.status_code != 200:
                                logger.error(
                                    f"Download failed for {image.path}: "
                                    f"status {response.status_code}"
                                )
                                error_count += 1
                                continue

                            os.makedirs(os.path.dirname(full_path), exist_ok=True)
                            with open(full_path, "wb") as f:
                                f.write(response.content)

                            if not image.dominant_color:
                                from backend.utils.colorExtractor import (
                                    extract_dominant_color,
                                )

                                color = await extract_dominant_color(full_path)
                                if color is not None:
                                    image.dominant_color = color
                                    await session.flush()

                            await session.commit()
                            logger.info(
                                f"Downloaded missing image: {image.path}, full path {full_path}, url {image.url}"
                            )
                            fixed_count += 1
                        except Exception as e:
                            logger.error(f"Error fixing image {image.path}: {e}")
                            error_count += 1

                    logger.info(
                        f"Fixed {fixed_count} images, "
                        f"skipped {skipped_count} (no url), {error_count} errors"
                    )

            elif command == "update-video-durations":
                from backend.constants import MEDIA_PATH
                from backend.youtube.access.videoAccess import VideoAccess

                try:
                    import ffmpeg  # type: ignore
                except ImportError:
                    logger.error(
                        "ffmpeg-python is required. Install with: pip install ffmpeg-python"
                    )
                    continue

                async with rockit_db.session_scope_async() as session:
                    a_result = await VideoAccess.get_all_videos_async(session=session)
                    if a_result.is_not_ok():
                        logger.error(f"Error getting videos: {a_result.message()}")
                        continue

                    videos = a_result.result()
                    logger.info(f"Found {len(videos)} videos")

                    updated_count = 0
                    error_count = 0

                    for video in videos:
                        if not video.video_path:
                            logger.warning(
                                f"Video {video.id} has no video_path, skipping"
                            )
                            error_count += 1
                            continue

                        full_path = os.path.join(MEDIA_PATH, video.video_path)
                        if not os.path.exists(full_path):
                            logger.warning(f"File not found: {full_path}, skipping")
                            error_count += 1
                            continue

                        try:
                            probe = ffmpeg.probe(full_path)  # type: ignore

                            video_info = next(  # type: ignore
                                s
                                for s in probe["streams"]  # type: ignore
                                if s["codec_type"] == "video"  # type: ignore
                            )
                            # type: ignore
                            duration = float(probe["format"]["duration"])  # type: ignore

                            real_duration = int(
                                float(video_info.get("duration", duration)) * 1000  # type: ignore
                            )

                            video.duration_ms = int(duration * 1000)
                            video.real_duration_ms = real_duration

                            await session.commit()
                            logger.info(
                                f"Updated video {video.id}: duration={video.duration_ms}, "
                                f"real_duration={video.real_duration_ms}"
                            )
                            updated_count += 1
                        except Exception as e:
                            logger.error(f"Error processing {video.id}: {e}")
                            error_count += 1

                    logger.info(f"Updated {updated_count} videos, {error_count} errors")

            elif command.startswith("fix-shared-images"):
                parts = command.split()
                sqlite_db: str = args.db

                for i, part in enumerate(parts):
                    if part == "--db" and i + 1 < len(parts):
                        sqlite_db = parts[i + 1]

                await fix_shared_images_async(sqlite_path=sqlite_db)

            elif command.startswith("migrate-playlist"):
                from backend.migratePlaylist import migrate_playlist_async

                parts = command.split()
                sqlite_db = "database.db"
                playlist_id = ""
                user_id = 0

                for i, part in enumerate(parts):
                    if part == "--db" and i + 1 < len(parts):
                        sqlite_db = parts[i + 1]
                    elif part == "--playlist-id" and i + 1 < len(parts):
                        playlist_id = parts[i + 1]
                    elif part == "--user-id" and i + 1 < len(parts):
                        user_id = int(parts[i + 1])

                if not playlist_id or not user_id:
                    logger.error(
                        "Usage: migrate-playlist --playlist-id <id> --user-id <int> [--db <path>]"
                    )
                    continue

                await migrate_playlist_async(
                    sqlite_path=sqlite_db,
                    playlist_id=playlist_id,
                    target_user_id=user_id,
                )

            elif command == "import-lrc-lib":
                from backend.lrclib.framework.importLrcLib import (
                    import_lrc_lib_from_dump_async,
                )

                lrc_lib_path = args.lrc_lib_path
                if not os.path.exists(lrc_lib_path):
                    logger.error(f"LRCLIB SQLite dump not found: {lrc_lib_path}")
                    continue

                async with rockit_db.session_scope_async() as session:
                    await import_lrc_lib_from_dump_async(
                        sqlite_path=lrc_lib_path,
                        session=session,
                    )

            else:
                print("Command not found.")

            if command_to_run != "":
                break

    print("Bye!")


if __name__ == "__main__":
    import asyncio

    try:
        loop = asyncio.get_running_loop()
        loop.create_task(main())
    except RuntimeError:
        asyncio.run(main())
